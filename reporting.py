import json
import shutil
import openpyxl
from openpyxl.drawing.image import Image as XlsxImage
from datetime import datetime

class ReportMaker:

    def __init__(self, template_path):
        self.template_path = template_path

    @staticmethod
    def copy_file(src_path, dst_path):
        try:
            shutil.copy2(src_path, dst_path)
        except IOError as e:
            print(f"파일 복사 오류: {e}")

    def create_report(self, output_path):
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
            
        with open('conditions.json', 'r') as file:
            report = json.load(file)
        with open('calculation_raw.json', 'r') as file:
            report.update(json.load(file)["report"])

        # make float values into string
        for i, j in report.items():
            if isinstance(j, float):
                if i == "r^2-" or i == "r^2+":
                    report[i] = f"{j:.4f}"
                else:
                    report[i] = f"{j:.2f}"

        now = datetime.now().strftime("%d%m%Y-%H%M%S")
        with open(f'./reports/report_{now}.json', 'w') as file:
            json.dump(report, file, indent=4)

        for row in ws.iter_rows():
            for cell in row:
                try:
                    report_item = cell.value.split("*")
                    if item := report.get(report_item[1]):
                        cell.value = item
                    else:
                        cell.value = "-"
                except IndexError:
                    pass
                except AttributeError:
                    pass

        wb.save(output_path)
        wb.close()

    @staticmethod
    def insert_image(file_path, image_path, cell, width=None, height=None):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        img = XlsxImage(image_path)
        img.width = width if width else img.width
        img.height = height if height else img.height
        ws.add_image(img, cell)

        wb.save(file_path)
        wb.close()

    @staticmethod
    def protect_excel_file(file_path, key):
        wb = openpyxl.load_workbook(file_path)
        
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            ws.protection.sheet = True
            ws.protection.password = key
        
        wb.save(file_path)
        wb.close()
    
    # @staticmethod    
    # def xlsx_to_image(input_file, output_file):
    #     # Load the Excel workbook
    #     workbook = openpyxl.load_workbook(filename=input_file)
    #     # Select the first sheet or specify the sheet name if needed
    #     sheet = workbook.active
    #     # Convert the sheet to an image using pyvips
    #     image = pyvips.Image.new_from_array(sheet.iter_rows(values_only=True))
    #     # Save the image to a file
    #     image.write_to_file(output_file)

    def process_report(self, output_path, image_path, cell, width, height, key):

        self.copy_file(self.template_path, output_path)
        self.create_report(output_path)
        self.insert_image(output_path, image_path, cell, width, height)
        self.protect_excel_file(output_path, key)
        # self.xlsx_to_image(output_path, output_path.split(".")[0] + ".jpg")

if __name__ == "__main__":
    # 템플릿 파일 이름
    template_path = "report_template.xlsx"
    # 결과 파일 명
    now = datetime.now().strftime("%y%m%d-%H%M%S")
    output_path = f"report_{now}.xlsx"
    # 이미지 정보
    image_path = "graph.png"
    cell = "B28"
    width = 590
    height = 355
    # 파일 보호 비밀번호
    key = "asdf"
    report_maker = ReportMaker(template_path)
    report_maker.process_report(output_path, image_path, cell, width, height, key)

    import platform
    if platform.system() == "Linux":
        import subprocess as sub
        sub.call(f"libreoffice --headless --convert-to pdf {output_path}", shell=True)
        shutil.move(output_path.split('.')[0] + '.pdf', "report.pdf")
        sub.call(f"xpdf ./report.pdf", shell=True)
